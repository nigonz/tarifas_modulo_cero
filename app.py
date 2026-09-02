import streamlit as st
import pandas as pd
import numpy as np
import io
import gc
from decimal import Decimal, ROUND_DOWN
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="TTR_ARIA - Sistema Integral", layout="wide")

# ==============================================================================
# FUNCIONES COMPARTIDAS
# ==============================================================================
def parse_argentinian_float(val):
    if pd.isna(val) or str(val).strip() == "": return 0.0
    s = str(val).replace(',', '.')
    try: return float(s)
    except: return 0.0

def calcular_tarifa(base, mult_ts, mult_nom):
    try:
        res = Decimal(str(base)) * Decimal(str(mult_ts)) * Decimal(str(mult_nom))
        return float(res.quantize(Decimal('0.01'), rounding=ROUND_DOWN))
    except: return 0.0

def formatear_excel(ws, df, cols_moneda):
    encabezado = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    relleno = PatternFill('solid', start_color='1F4E78')
    cuerpo = Font(name='Arial', size=10)
    for celda in ws[1]:
        celda.font = encabezado
        celda.fill = relleno
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for fila in ws.iter_rows(min_row=2):
        for celda in fila: celda.font = cuerpo
    for i, col in enumerate(df.columns, start=1):
        letra = get_column_letter(i)
        ws.column_dimensions[letra].width = min(max(len(str(col)) + 4, 12), 32)
        if col in cols_moneda:
            for celda in ws[letra][1:]: celda.number_format = '#,##0.00'
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

def resumir(df, claves, agg_estandar, **extra):
    claves_limpias = [c for c in claves if c] # Elimina Nones por seguridad
    return df.groupby(claves_limpias, dropna=False).agg(**agg_estandar, **extra).reset_index()

def buscar_col(df, *palabras_clave):
    for kw in palabras_clave:
        for col in df.columns:
            if str(col).strip().upper() == kw.upper(): return col
    for kw in palabras_clave:
        for col in df.columns:
            if kw.upper() in str(col).upper(): return col
    return None

# ==============================================================================
# MÓDULO 0: TARIFARIO JN
# ==============================================================================
def modulo_tarifas():
    st.title("🚜 TTR_ARIA - Módulo 0: Tarifario JN")
    st.markdown("Generador de cuadros tarifarios con indexación automática y arquitectura de espejos.")
    col_izq, col_der = st.columns([1, 3])
    with col_izq:
        archivo_historico = st.file_uploader("Subir Matriz Histórica (.xlsx)", type=['xlsx'], key='mod0_hist')
        mes_act = st.text_input("Nombre del Mes Nuevo", "Septiembre")
    with col_der:
        st.info("💡 Ingresá tus 10 bases manuales. La base 1-4KMCN se indexará automáticamente.")
        llaves = ['1SCN', '2SCN', '3SCN', '4SCN', '5SCN', '5KPCN', '6KPCN', '7KPCN', '8KPCN', '9KPCN']
        if 'tabla_bases' not in st.session_state:
            st.session_state.tabla_bases = pd.DataFrame({
                "CONCAT Base": llaves, "Límite Inferior": pd.Series([""] * 10, dtype=str), "Límite Superior": pd.Series([""] * 10, dtype=str)
            })
        tarifas_editadas = st.data_editor(st.session_state.tabla_bases, hide_index=True, use_container_width=True)
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("🪄 Auto-Completar Límites Superiores", use_container_width=True):
                df_temp = tarifas_editadas.copy()
                for i in range(5):
                    if str(df_temp.loc[i, "Límite Inferior"]).strip() != "": df_temp.loc[i, "Límite Superior"] = df_temp.loc[i, "Límite Inferior"]
                for i in range(5, 9):
                    if str(df_temp.loc[i+1, "Límite Inferior"]).strip() != "": df_temp.loc[i, "Límite Superior"] = df_temp.loc[i+1, "Límite Inferior"]
                st.session_state.tabla_bases = df_temp
                st.rerun()
        with col_btn2:
            if st.button("Generar Cuadro Tarifario TTR", type="primary", use_container_width=True):
                dict_bases_inf = {k: parse_argentinian_float(v) for k, v in zip(tarifas_editadas['CONCAT Base'], tarifas_editadas['Límite Inferior'])}
                dict_bases_sup = {k: parse_argentinian_float(v) for k, v in zip(tarifas_editadas['CONCAT Base'], tarifas_editadas['Límite Superior'])}
                if archivo_historico is None: st.error("⚠️ Subí la matriz histórica.")
                elif all(v == 0.0 for v in dict_bases_inf.values()): st.error("⚠️ Faltan cargar bases.")
                else:
                    with st.spinner("Procesando matriz..."):
                        try:
                            df_hist = pd.read_excel(archivo_historico, header=0)
                            c_concat = buscar_col(df_hist, 'concat'); c_ts = buscar_col(df_hist, 'ts')
                            c_nom = buscar_col(df_hist, 'nominaliz'); c_km = buscar_col(df_hist, 'km')

                            columnas_protegidas = [c_concat, c_ts, c_nom, c_km, 'Seccion', 'TIPO SECCION']
                            cols_historicas_meses = []
                            for col in df_hist.columns:
                                if col not in columnas_protegidas and not str(col).startswith('Unnamed'):
                                    df_hist[col] = pd.to_numeric(df_hist[col], errors='coerce').round(2)
                                    cols_historicas_meses.append(col)

                            base_1_4kmcn_dinamica = 0.0
                            if len(cols_historicas_meses) > 0:
                                col_mes_anterior = cols_historicas_meses[0]
                                idx_1scn = df_hist[df_hist[c_concat].astype(str).str.strip().str.upper() == '1SCN'].index
                                idx_1_4kmcn = df_hist[df_hist[c_concat].astype(str).str.strip().str.upper() == '1-4KMCN'].index
                                if len(idx_1scn) > 0 and len(idx_1_4kmcn) > 0:
                                    val_1scn_viejo = df_hist.loc[idx_1scn, col_mes_anterior].mode()[0]
                                    val_1_4kmcn_viejo = df_hist.loc[idx_1_4kmcn, col_mes_anterior].mode()[0]
                                    val_1scn_nuevo = dict_bases_inf.get('1SCN', 0.0)
                                    factor_aumento = (val_1scn_nuevo / float(val_1scn_viejo)) if (pd.notna(val_1scn_viejo) and val_1scn_viejo != 0) else 1.0
                                    if pd.notna(val_1_4kmcn_viejo): base_1_4kmcn_dinamica = float(val_1_4kmcn_viejo) * factor_aumento

                            nuevos_limites_inf, nuevos_limites_sup = [], []
                            mapa_resultados = {}
                            for _, row in df_hist.iterrows():
                                concat = str(row[c_concat]).strip().upper()
                                if concat in ['NAN', 'NONE', ''] or "SR" in concat:
                                    nuevos_limites_inf.append(np.nan); nuevos_limites_sup.append(np.nan); continue
                                ts = str(row[c_ts]).strip().upper(); nom = str(row[c_nom]).strip().upper()
                                km_str = str(row.get(c_km, '')).strip()
                                es_caso_especial_km2 = (concat.startswith("1-4KM") and "2" in concat)

                                if es_caso_especial_km2:
                                    mapa_km2 = {'45-60':('1-4KMCN','5KPCN'), '60-75':('1-4KMEN','5KPEN'), '75-90':('1-4KMEAN','5KPEAN'),
                                                '90-150':('1-4KMCSN','5KPCSN'), '0-3':('1-4KMESN','5KPESN'), '3-6':('1-4KMEASN','5KPEASN')}
                                    ref_inf, ref_sup = mapa_km2.get(km_str, ('1-4KMCN','5KPCN'))
                                    val_inf, val_sup = mapa_resultados.get(ref_inf, (0,0))[0], mapa_resultados.get(ref_sup, (0,0))[1]
                                else:
                                    base_key_inf = base_key_sup = "1SCN"
                                    if concat.startswith("1S"): base_key_inf = base_key_sup = "1SCN"
                                    elif concat.startswith("2S"): base_key_inf = base_key_sup = "2SCN"
                                    elif concat.startswith("3S"): base_key_inf = base_key_sup = "3SCN"
                                    elif concat.startswith("4S"): base_key_inf = base_key_sup = "4SCN"
                                    elif concat.startswith("5S"): base_key_inf = base_key_sup = "5SCN"
                                    elif concat.startswith("1-4KM"): base_key_inf = base_key_sup = "1-4KMCN"
                                    elif concat.startswith("5KP"): base_key_inf = base_key_sup = "5KPCN"
                                    elif concat.startswith("6KP"): base_key_inf = base_key_sup = "6KPCN"
                                    elif concat.startswith("7KP"): base_key_inf = base_key_sup = "7KPCN"
                                    elif concat.startswith("8KP"): base_key_inf = base_key_sup = "8KPCN"
                                    elif concat.startswith("9KP"): base_key_inf = base_key_sup = "9KPCN"

                                    if base_key_inf == "1-4KMCN": base_inf = base_sup = base_1_4kmcn_dinamica
                                    else:
                                        base_inf = float(dict_bases_inf.get(base_key_inf, 0)); base_sup = float(dict_bases_sup.get(base_key_sup, 0))
                                    if concat.startswith("5KP") and ts == "E": base_sup = base_inf
                                    mult_ts = 1.75 if ts == "EA" else (1.25 if ts == "E" else 1.0)
                                    mult_nom = 2.0 if "SN" in nom else 1.0

                                    val_inf = calcular_tarifa(base_inf, mult_ts, mult_nom)
                                    val_sup = calcular_tarifa(base_sup, mult_ts, mult_nom)
                                    if concat not in mapa_resultados: mapa_resultados[concat] = (val_inf, val_sup)

                                nuevos_limites_inf.append(val_inf); nuevos_limites_sup.append(val_sup)

                            df_hist[f'{mes_act}'] = nuevos_limites_inf
                            col_sup_name = f'{mes_act}_Sup'
                            df_hist[col_sup_name] = nuevos_limites_sup
                            rename_dict = {}
                            espacios = 1
                            for col in df_hist.columns:
                                if "Unnamed" in str(col) or str(col) == col_sup_name:
                                    rename_dict[col] = " " * espacios; espacios += 1
                            df_export = df_hist.rename(columns=rename_dict)

                            buffer = io.BytesIO()
                            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                                df_export.to_excel(writer, index=False, sheet_name='Matriz_ARIA')
                            buffer.seek(0)
                            wb = openpyxl.load_workbook(buffer)
                            ws = wb.active
                            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                                for cell in row:
                                    if isinstance(cell.value, (int, float)): cell.number_format = '0.00'
                            final_buffer = io.BytesIO()
                            wb.save(final_buffer); final_buffer.seek(0)

                            st.success(f"✅ ¡Matriz de {mes_act} generada con éxito!")
                            st.download_button(label=f"📥 Descargar Matriz Definitiva {mes_act} (.xlsx)", data=final_buffer.getvalue(), file_name=f"Matriz_TTR_ARIA_{mes_act}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        except Exception as e: st.error(f"Error: {e}")

# ==============================================================================
# MÓDULO 1: DMK (ITG y ATS) - VERSIÓN OPTIMIZADA EN MEMORIA (OOM KILLER SAFE)
# ==============================================================================
def modulo_dmk():
    st.title("🧮 TTR_ARIA - Módulo 1: Liquidación DMK (ITG/ATS)")
    st.markdown("Pipeline integral de enriquecimiento, control y generación de reportes de compensaciones.")

    with st.expander("⚙️ Parámetros de Negocio", expanded=False):
        col1, col2, col3 = st.columns(3)
        IVA = col1.number_input("Factor IVA (Ej: 1.105)", value=1.105, format="%.3f")
        str_contratos_ats = col2.text_input("Contratos ATS (separados por coma)", "621")
        str_contratos_est = col3.text_input("Contratos Estudiantiles", "830, 831, 832, 833")
        
        CONTRATOS_ATS = [int(x.strip()) for x in str_contratos_ats.split(',')]
        CONTRATOS_ESTUDIANTILES = [int(x.strip()) for x in str_contratos_est.split(',')]
        ENERGIA_DEFECTO = 3
        GRUPOS_AMBA = ['DF', 'SGI', 'SGII', 'SGIKM', 'UMA1', 'UMA2', 'UPA', 'UPAKM']
        GRUPO_INP = 'INP'
        SIN_DATO = 'S/D'

    st.header("1. Carga de Archivos de Entrada")
    col1, col2 = st.columns(2)
    file_dggi = col1.file_uploader("1. Base DGGI (CSV o Excel)", type=['csv', 'xlsx'])
    file_nom_univ = col2.file_uploader("2. Nomenclador Líneas (07. Nomenclador)", type=['xlsx'])
    file_nom_ramal = col1.file_uploader("3. Nomenclador Ramal - TS", type=['xlsx'])
    file_pme = col2.file_uploader("4. Parque Móvil - Energías", type=['xlsx'])

    if st.button("🚀 Ejecutar Liquidación DMK", type="primary", use_container_width=True):
        if not all([file_dggi, file_nom_univ, file_nom_ramal, file_pme]):
            st.error("⚠️ Faltan cargar archivos. Asegurate de subir los 4 requeridos.")
            return
            
        with st.spinner("Procesando pipeline de datos DMK (Optimizando RAM)..."):
            try:
                # 1. LECTURA DE NOMENCLADORES
                xls_nom = pd.ExcelFile(file_nom_univ)
                hoja_nom = 'Nomenclador_Interior' if 'Nomenclador_Interior' in xls_nom.sheet_names else xls_nom.sheet_names[0]
                nom_lineas_raw = pd.read_excel(xls_nom, sheet_name=hoja_nom)
                
                xls_ram = pd.ExcelFile(file_nom_ramal)
                hoja_ram = 'NOMENCLADOR TS' if 'NOMENCLADOR TS' in xls_ram.sheet_names else xls_ram.sheet_names[0]
                nom_ramal_raw = pd.read_excel(xls_ram, sheet_name=hoja_ram)
                
                pme_raw = pd.read_excel(file_pme, sheet_name='Nomenclador_PM_E')
                pme_raw.columns = pme_raw.columns.str.strip().str.upper().str.replace('Í', 'I').str.replace('É', 'E')
                tipo_energia_raw = pd.read_excel(file_pme, sheet_name='Tipo_Energia')
                tipo_energia_raw.columns = tipo_energia_raw.columns.str.strip().str.upper().str.replace('Í', 'I').str.replace('É', 'E')

                # 2. PROCESAMIENTO DE NOMENCLADORES
                n_lin = pd.DataFrame()
                c_id_linea = buscar_col(nom_lineas_raw, 'ID_LINEA', 'ID LINEA')
                c_gt = buscar_col(nom_lineas_raw, 'GT', 'GRUPO_TARIFARIO', 'GRUPO TARIFARIO')
                c_silas = buscar_col(nom_lineas_raw, 'SILAS - AMBA', 'LINEA_SILAS_DNGFF', 'LINEA SILAS DNGFF', 'SILAS')
                c_empresa = buscar_col(nom_lineas_raw, 'ID_EMPRESA', 'IDEMPRESA', 'EMPRESA')
                c_rs = buscar_col(nom_lineas_raw, 'Razon social', 'RAZON_SOCIAL', 'SOCIAL')
                c_juris = buscar_col(nom_lineas_raw, 'Jurisdiccion', 'JURIS')
                c_prov = buscar_col(nom_lineas_raw, 'Provincia', 'PROV')
                c_mun = buscar_col(nom_lineas_raw, 'Localidad', 'MUNICIPIO')
                c_depto = buscar_col(nom_lineas_raw, 'Departamento', 'DEPTO')

                n_lin['ID_LINEA'] = pd.to_numeric(nom_lineas_raw[c_id_linea], errors='coerce').astype('Int64') if c_id_linea else pd.Series(dtype='Int64')
                n_lin['GRUPO_TARIFARIO'] = nom_lineas_raw[c_gt].astype('string').str.strip() if c_gt else pd.Series(dtype='string')
                n_lin['LINEA_SILAS_DNGFF'] = nom_lineas_raw[c_silas].astype('string').str.strip() if c_silas else pd.Series(dtype='string')
                n_lin['ID_EMPRESA_NOM'] = pd.to_numeric(nom_lineas_raw[c_empresa], errors='coerce').astype('Int64') if c_empresa else pd.Series(dtype='Int64')
                n_lin['RAZON_SOCIAL'] = nom_lineas_raw[c_rs].astype('string').str.strip() if c_rs else pd.Series(dtype='string')
                n_lin['JURISDICCION'] = nom_lineas_raw[c_juris].astype('string').str.strip() if c_juris else pd.Series(dtype='string')
                n_lin['PROVINCIA'] = nom_lineas_raw[c_prov].astype('string').str.strip() if c_prov else pd.Series(dtype='string')
                n_lin['MUNICIPIO'] = nom_lineas_raw[c_mun].astype('string').str.strip() if c_mun else pd.Series(dtype='string')
                n_lin['DEPARTAMENTO'] = nom_lineas_raw[c_depto].astype('string').str.strip() if c_depto else SIN_DATO
                nom_lineas = n_lin.dropna(subset=['ID_LINEA']).drop_duplicates(subset=['ID_LINEA'])

                c_ramal = buscar_col(nom_ramal_raw, 'IdRamalNS', 'RAMAL')
                c_ts = buscar_col(nom_ramal_raw, 'TIPO DE SERVICIO FINAL', 'SERVICIO', 'TS')
                n_ram = pd.DataFrame()
                n_ram['RAMAL'] = pd.to_numeric(nom_ramal_raw[c_ramal], errors='coerce').astype('Int64') if c_ramal else pd.Series(dtype='Int64')
                n_ram['TIPO_SERVICIO'] = nom_ramal_raw[c_ts].astype('string').str.strip().str.upper() if c_ts else pd.Series(dtype='string')
                nom_ramal = n_ram.dropna(subset=['RAMAL']).drop_duplicates(subset=['RAMAL'])

                c_dom_pm = buscar_col(pme_raw, 'DOMINIO', 'PATENTE')
                c_ene_pm = buscar_col(pme_raw, 'ENERGIA', 'ENERG')
                p = pd.DataFrame()
                p['DOMINIO'] = pme_raw[c_dom_pm].astype('string').str.strip().str.upper() if c_dom_pm else pd.Series(dtype='string')
                p['ENERGIA_PM'] = pd.to_numeric(pme_raw[c_ene_pm], errors='coerce').astype('Int64') if c_ene_pm else pd.Series(dtype='Int64')
                pme = p.dropna(subset=['DOMINIO']).drop_duplicates(subset=['DOMINIO'])

                c_te_ene = buscar_col(tipo_energia_raw, 'ENERGIA', 'ENERG', 'ID')
                c_te_con = buscar_col(tipo_energia_raw, 'CONCEPTO', 'CONCEPT', 'DESC')
                if c_te_ene and c_te_con:
                    MAPA_ENERGIA = dict(zip(pd.to_numeric(tipo_energia_raw[c_te_ene], errors='coerce'), tipo_energia_raw[c_te_con].astype(str).str.strip().str.upper()))
                else: MAPA_ENERGIA = {}

                # Destruimos los Raws de memoria
                del nom_lineas_raw, nom_ramal_raw, pme_raw, tipo_energia_raw, n_lin, n_ram, p
                gc.collect()

                # 3. LECTURA Y ESTANDARIZACIÓN IN-PLACE DE DGGI (RAM Saver)
                d = pd.read_csv(file_dggi, encoding='ISO-8859-1', delimiter=';') if file_dggi.name.endswith('.csv') else pd.read_excel(file_dggi)
                d.columns = d.columns.str.strip()
                
                c_monto = buscar_col(d, 'MONTO', 'RECAUDACION')
                if c_monto: d.rename(columns={c_monto: 'RECAUDACION'}, inplace=True)
                
                for c in ['DOMINIO', 'MK', 'VIAJE INTEGRADO', 'MEDIOS_DE_PAGO']:
                    col_real = buscar_col(d, c)
                    if col_real: d[col_real] = d[col_real].astype('string').str.strip().str.upper()
                for c in ['TARIFA', 'DEBITADO', 'DESCUENTO X INTEGRACION', 'CANTIDAD_USOS', 'RECAUDACION', 'TOTAL DESC POR INTEGRACION', 'DESCUENTO_TOTAL', 'DESCUENTO_ATRIBUTOS']:
                    col_real = buscar_col(d, c)
                    if col_real: d[col_real] = pd.to_numeric(d[col_real], errors='coerce').fillna(0)
                for c in ['ID_EMPRESA', 'ID_LINEA', 'RAMAL', 'CONTRATO', 'INTERNO']:
                    col_real = buscar_col(d, c)
                    if col_real: d[col_real] = pd.to_numeric(d[col_real], errors='coerce').astype('Int64')

                # 4. FUSIONES (MERGES)
                d = d.merge(nom_lineas, on='ID_LINEA', how='left', validate='m:1')
                d = d.merge(nom_ramal, on='RAMAL', how='left', validate='m:1')
                d = d.merge(pme, on='DOMINIO', how='left', validate='m:1')
                
                del nom_lineas, nom_ramal, pme
                gc.collect()
                
                # 5. CÁLCULOS
                d['ES_BENEFICIARIA'] = np.where(d['GRUPO_TARIFARIO'].notna(), 'SI', 'NO')
                gt_upper = d['GRUPO_TARIFARIO'].astype('string').str.strip().str.upper()
                d['AMBA'] = np.select([gt_upper.isin(GRUPOS_AMBA).fillna(False).to_numpy(), gt_upper.eq(GRUPO_INP).fillna(False).to_numpy()], ['SI', 'AMBA - INP'], default='NO')
                
                d['EN_PARQUE_MOVIL'] = np.where(d['ENERGIA_PM'].notna(), 'SI', 'NO')
                d['TIPO_ENERGIA'] = d['ENERGIA_PM'].fillna(ENERGIA_DEFECTO).astype('Int64')
                d['ENERGIA_DESC'] = d['TIPO_ENERGIA'].map(MAPA_ENERGIA).astype('string')
                d.drop(columns=['ENERGIA_PM'], inplace=True, errors='ignore')
                
                for c in ['JURISDICCION', 'PROVINCIA', 'MUNICIPIO', 'DEPARTAMENTO', 'GRUPO_TARIFARIO', 'RAZON_SOCIAL', 'LINEA_SILAS_DNGFF', 'TIPO_SERVICIO']:
                    d[c] = d[c].astype('string').fillna(SIN_DATO)
                
                d['COINCIDE_EMPRESA'] = np.select([d['ID_EMPRESA_NOM'].isna().to_numpy(), (d['ID_EMPRESA'] == d['ID_EMPRESA_NOM']).fillna(False).to_numpy()], [SIN_DATO, 'SI'], default='NO')

                col_viaje_int = buscar_col(d, 'VIAJE INTEGRADO')
                d['ES_INTEGRADO'] = np.where(d[col_viaje_int].astype('string').eq('SI') if col_viaje_int else False, 'SI', 'NO')
                
                es_ats = d['CONTRATO'].isin(CONTRATOS_ATS).to_numpy()
                es_est = d['CONTRATO'].isin(CONTRATOS_ESTUDIANTILES).to_numpy()
                col_desc_atri = buscar_col(d, 'DESCUENTO_ATRIBUTOS')
                tiene_desc = (d[col_desc_atri] > 0).to_numpy() if col_desc_atri else np.zeros(len(d), dtype=bool)
                
                d['ES_ATS'] = np.where(es_ats, 'SI', 'NO')
                d['ES_ESTUDIANTIL'] = np.where(es_est, 'SI', 'NO')
                d['ES_OTRO_BENEFICIO'] = np.where(tiene_desc & ~es_ats & ~es_est, 'SI', 'NO')
                d['TIPO_BENEFICIO'] = np.select([es_ats, es_est, tiene_desc & ~es_ats & ~es_est], ['ATS', 'ESTUDIANTIL', 'OTRO BENEFICIO'], default='SIN BENEFICIO')

                c_usos = buscar_col(d, 'CANTIDAD_USOS')
                c_total_desc_int = buscar_col(d, 'TOTAL DESC POR INTEGRACION')
                c_desc_tot = buscar_col(d, 'DESCUENTO_TOTAL')
                c_deb = buscar_col(d, 'DEBITADO')
                c_tar = buscar_col(d, 'TARIFA')
                c_desc_x_int = buscar_col(d, 'DESCUENTO X INTEGRACION')

                u = d[c_usos] if c_usos else 0
                d['COMP. ITG'] = d[c_total_desc_int] if c_total_desc_int else 0
                d['COMP. ITG s/IVA'] = d['COMP. ITG'] / IVA
                d['COMP. ATS'] = np.where(es_ats, d[col_desc_atri] if col_desc_atri else 0, 0.0)
                d['COMP. ATS s/IVA'] = d['COMP. ATS'] / IVA
                d['DESCUENTO_TOTAL s/IVA'] = (d[c_desc_tot] if c_desc_tot else 0) / IVA
                d['COMP. TOTAL s/IVA'] = d['COMP. ITG s/IVA'] + d['COMP. ATS s/IVA']
                
                d['RECAUDACION_CALC'] = (d[c_deb] if c_deb else 0) * u
                d['DESC_TOTAL_CALC'] = ((d[c_tar] if c_tar else 0) - (d[c_deb] if c_deb else 0)) * u
                d['COMP_ITG_CALC'] = (d[c_desc_x_int] if c_desc_x_int else 0) * u
                d['COMP_ATS_CALC'] = np.where(es_ats, ((d[c_tar] if c_tar else 0) - (d[c_deb] if c_deb else 0) - (d[c_desc_x_int] if c_desc_x_int else 0)) * u, 0.0)
                d['DIF_RECAUDACION'] = (d['RECAUDACION_CALC'] - d['RECAUDACION']).round(2)
                d['DIF_DESC_TOTAL'] = (d['DESC_TOTAL_CALC'] - (d[c_desc_tot] if c_desc_tot else 0)).round(2)
                d['DIF_ITG'] = (d['COMP_ITG_CALC'] - d['COMP. ITG']).round(2)
                d['DIF_ATS'] = (d['COMP_ATS_CALC'] - d['COMP. ATS']).round(2)

                # 6. SEPARACIÓN EN BENEFICIARIAS Y ORDEN
                df_final = d[d['ES_BENEFICIARIA'] == 'SI']
                df_no_benef = d[d['ES_BENEFICIARIA'] == 'NO']

                ORDEN_BASE = [
                    'ID_EMPRESA', 'ID_EMPRESA_NOM', 'COINCIDE_EMPRESA', 'RAZON_SOCIAL', 'ID_LINEA', 'LINEA_SILAS_DNGFF', 'RAMAL', 'TIPO_SERVICIO', 'INTERNO', 'DOMINIO', 'MK',
                    'JURISDICCION', 'PROVINCIA', 'MUNICIPIO', 'DEPARTAMENTO', 'GRUPO_TARIFARIO', 'AMBA', 'ES_BENEFICIARIA',
                    'EN_PARQUE_MOVIL', 'TIPO_ENERGIA', 'ENERGIA_DESC',
                    'CONTRATO', 'MEDIOS_DE_PAGO', 'VIAJE INTEGRADO', 'ES_INTEGRADO', 'TIPO_BENEFICIO', 'ES_ATS', 'ES_ESTUDIANTIL', 'ES_OTRO_BENEFICIO',
                    'TARIFA', 'DEBITADO', 'DESCUENTO X INTEGRACION', 'CANTIDAD_USOS',
                    'RECAUDACION', 'DESCUENTO_TOTAL', 'DESCUENTO_TOTAL s/IVA', 'TOTAL DESC POR INTEGRACION', 'DESCUENTO_ATRIBUTOS',
                    'COMP. ITG', 'COMP. ITG s/IVA', 'COMP. ATS', 'COMP. ATS s/IVA', 'COMP. TOTAL s/IVA',
                    'RECAUDACION_CALC', 'DESC_TOTAL_CALC', 'COMP_ITG_CALC', 'COMP_ATS_CALC', 'DIF_RECAUDACION', 'DIF_DESC_TOTAL', 'DIF_ITG', 'DIF_ATS',
                ]
                
                df_final = df_final[[c for c in ORDEN_BASE if c in df_final.columns] + [c for c in df_final.columns if c not in ORDEN_BASE]]
                df_no_benef = df_no_benef[[c for c in ORDEN_BASE if c in df_no_benef.columns] + [c for c in df_no_benef.columns if c not in ORDEN_BASE]]

                # Destruimos "d" principal para liberar memoria masiva
                del d
                gc.collect()

                # 7. RESÚMENES
                c_us_df = buscar_col(df_final, 'CANTIDAD_USOS')
                c_dt_df = buscar_col(df_final, 'DESCUENTO_TOTAL')
                c_da_df = buscar_col(df_final, 'DESCUENTO_ATRIBUTOS')
                c_di_df = buscar_col(df_final, 'TOTAL DESC POR INTEGRACION')
                
                AGG_ESTANDAR = dict(RECAUDACION=('RECAUDACION', 'sum'), USOS=(c_us_df, 'sum'), DESCUENTO_TOTAL=(c_dt_df, 'sum'), DESCUENTO_TOTAL_sIVA=('DESCUENTO_TOTAL s/IVA', 'sum'), COMP_ITG=('COMP. ITG', 'sum'), COMP_ITG_sIVA=('COMP. ITG s/IVA', 'sum'), COMP_ATS=('COMP. ATS', 'sum'), COMP_ATS_sIVA=('COMP. ATS s/IVA', 'sum'))
                
                resumen_compensacion = resumir(df_final, ['JURISDICCION', 'PROVINCIA', 'MUNICIPIO', 'GRUPO_TARIFARIO', 'AMBA'], AGG_ESTANDAR)
                resumen_unicos = df_final.groupby(['PROVINCIA', 'GRUPO_TARIFARIO', 'AMBA'], dropna=False).agg(LINEAS_SILAS_UNICAS=('LINEA_SILAS_DNGFF', 'nunique'), ID_LINEAS_UNICAS=('ID_LINEA', 'nunique'), RAMALES_UNICOS=('RAMAL', 'nunique'), EMPRESAS_UNICAS=('ID_EMPRESA', 'nunique'), INTERNOS_UNICOS=('INTERNO', 'nunique'), DOMINIOS_UNICOS=('DOMINIO', 'nunique')).reset_index()
                resumen_energia = resumir(df_final[df_final['EN_PARQUE_MOVIL'] == 'SI'], ['ID_LINEA', 'LINEA_SILAS_DNGFF', 'PROVINCIA', 'AMBA', 'ENERGIA_DESC'], AGG_ESTANDAR)
                resumen_contrato = df_final.groupby(['CONTRATO', 'AMBA'], dropna=False).agg(RECAUDACION=('RECAUDACION', 'sum'), USOS=(c_us_df, 'sum'), DESCUENTO_TOTAL_sIVA=('DESCUENTO_TOTAL s/IVA', 'sum'), COMP_ITG_sIVA=('COMP. ITG s/IVA', 'sum'), COMP_ATS_sIVA=('COMP. ATS s/IVA', 'sum'), ATRIBUTO_EN_BASE=(c_da_df, 'sum')).reset_index()
                
                c_mp = buscar_col(df_final, 'MEDIOS_DE_PAGO')
                resumen_medio_pago = resumir(df_final, [c_mp, 'PROVINCIA', 'MUNICIPIO', 'GRUPO_TARIFARIO', 'AMBA'], AGG_ESTANDAR)
                
                control_cobertura = df_final.groupby(['ES_BENEFICIARIA', 'JURISDICCION', 'AMBA'], dropna=False).agg(**AGG_ESTANDAR, LINEAS=('ID_LINEA', 'nunique')).reset_index()

                CLAVES_TARIFARIO = ['PROVINCIA', 'MUNICIPIO', 'GRUPO_TARIFARIO', 'AMBA', 'LINEA_SILAS_DNGFF', 'ID_LINEA', 'RAMAL', 'CONTRATO', 'TARIFA', 'DEBITADO']
                resumen_tarifario = df_final.groupby([c for c in CLAVES_TARIFARIO if c in df_final.columns], dropna=False).agg(USOS=(c_us_df, 'sum'), RECAUDACION=('RECAUDACION', 'sum'), **{'TOTAL DESC POR INTEGRACION': (c_di_df, 'sum')}, DESCUENTO_ATRIBUTOS=(c_da_df, 'sum'), DESCUENTO_TOTAL=(c_dt_df, 'sum')).reset_index()

                def _listar(s): return ', '.join(str(v) for v in sorted(s.dropna().unique()))
                no_benef_detalle = df_no_benef.groupby(['ID_LINEA', 'ID_EMPRESA'], dropna=False).agg(RAMALES=('RAMAL', _listar), CONTRATOS=('CONTRATO', _listar), RECAUDACION=('RECAUDACION', 'sum'), USOS=(c_us_df, 'sum'), DESCUENTO_TOTAL_sIVA=('DESCUENTO_TOTAL s/IVA', 'sum'), COMP_ITG_sIVA=('COMP. ITG s/IVA', 'sum'), COMP_ATS_sIVA=('COMP. ATS s/IVA', 'sum'), ATRIBUTO_EN_BASE=(c_da_df, 'sum')).reset_index().sort_values('USOS', ascending=False)

                df_tarifario_dominio = df_final.copy()
                es_gasoil = (df_tarifario_dominio['TIPO_ENERGIA'] == 3).fillna(False).to_numpy()
                df_tarifario_dominio['DOMINIO'] = np.where(es_gasoil, 'NO', df_tarifario_dominio['DOMINIO'])
                df_tarifario_dominio['ENERGIA'] = df_tarifario_dominio['TIPO_ENERGIA']
                resumen_tarifario_dominio = df_tarifario_dominio.groupby([c for c in CLAVES_TARIFARIO + ['DOMINIO', 'ENERGIA'] if c in df_tarifario_dominio.columns], dropna=False).agg(USOS=(c_us_df, 'sum'), RECAUDACION=('RECAUDACION', 'sum'), **{'TOTAL DESC POR INTEGRACION': (c_di_df, 'sum')}, DESCUENTO_ATRIBUTOS=(c_da_df, 'sum'), DESCUENTO_TOTAL=(c_dt_df, 'sum')).reset_index()

                COLS_MONEDA = {
                    'TARIFA', 'DEBITADO', 'DESCUENTO X INTEGRACION', 'RECAUDACION', 'RECAUDACION_CALC', 'DESCUENTO_TOTAL', 'DESCUENTO_TOTAL s/IVA', 'DESCUENTO_TOTAL_sIVA', 'DESC_TOTAL_CALC',
                    'TOTAL DESC POR INTEGRACION', 'DESCUENTO_ATRIBUTOS', 'ATRIBUTO_EN_BASE', 'COMP. ITG', 'COMP. ITG s/IVA', 'COMP. ATS', 'COMP. ATS s/IVA', 'COMP. TOTAL s/IVA',
                    'COMP_ITG', 'COMP_ITG_sIVA', 'COMP_ATS', 'COMP_ATS_sIVA', 'COMP_ITG_CALC', 'COMP_ATS_CALC', 'DIF_RECAUDACION', 'DIF_DESC_TOTAL', 'DIF_ITG', 'DIF_ATS', 'MONTO TOTAL COBRADO', 'DESCUENTO TOTAL ITG', 'DESCUENTO TOTAL ATS'
                }

                base_621 = df_final[df_final['CONTRATO'].isin(CONTRATOS_ATS)]
                resumen_621 = base_621.groupby(['JURISDICCION', 'PROVINCIA', 'MUNICIPIO', 'ID_EMPRESA', 'RAZON_SOCIAL', 'ID_LINEA', 'RAMAL'], dropna=False).agg(**{'MONTO TOTAL COBRADO': ('RECAUDACION', 'sum')}, **{'CANTIDAD DE TRANSACCIONES': (c_us_df, 'sum')}, **{'DESCUENTO TOTAL ITG': ('COMP. ITG', 'sum')}, **{'DESCUENTO TOTAL ATS': ('COMP. ATS', 'sum')}, AMBA=('AMBA', 'first')).reset_index().rename(columns={'ID_LINEA': 'LINEA', 'RAZON_SOCIAL': 'RAZON SOCIAL'})

                # Exportación en buffers
                buf_resumenes = io.BytesIO()
                with pd.ExcelWriter(buf_resumenes, engine='openpyxl') as writer:
                    for nombre, tabla in {
                        'Resumen_Compensacion': resumen_compensacion, 'Resumen_Unicos': resumen_unicos, 'Resumen_Energia': resumen_energia,
                        'Resumen_Contrato': resumen_contrato, 'Resumen_MedioPago': resumen_medio_pago, 'Control_Cobertura': control_cobertura,
                        'Resumen_Tarifario': resumen_tarifario, 'Resumen_Tarifario_Dominio': resumen_tarifario_dominio, 'NoBenef_Detalle': no_benef_detalle
                    }.items():
                        tabla.to_excel(writer, index=False, sheet_name=nombre)
                        formatear_excel(writer.sheets[nombre], tabla, COLS_MONEDA)
                buf_resumenes.seek(0)

                buf_621 = io.BytesIO()
                with pd.ExcelWriter(buf_621, engine='openpyxl') as writer:
                    resumen_621.to_excel(writer, index=False, sheet_name='ATS_621')
                    formatear_excel(writer.sheets['ATS_621'], resumen_621, COLS_MONEDA)
                buf_621.seek(0)

                buf_csv = io.BytesIO()
                df_final.to_csv(buf_csv, index=False, sep=';', encoding='utf-8-sig')
                buf_csv.seek(0)

                # Liberación total final
                del df_final, df_no_benef, base_621
                gc.collect()

                st.success("✅ ¡Liquidación procesada con éxito sin agotar la memoria!")
                
                st.markdown("### Descargas Disponibles")
                d1, d2, d3 = st.columns(3)
                d1.download_button("📥 Descargar Resúmenes (.xlsx)", data=buf_resumenes, file_name="DGGI_ITG_ATS_Resumenes.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                d2.download_button("📥 Descargar Reporte ATS 621 (.xlsx)", data=buf_621, file_name="DGGI_ATS_621.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                d3.download_button("📥 Descargar Base Detalle (.csv)", data=buf_csv, file_name="DGGI_Base_Detalle.csv", mime="text/csv", use_container_width=True)

            except Exception as e: st.error(f"Error procesando DMK: {e}")

# ==============================================================================
# MÓDULO 3: CÁLCULO TTR (MOTOR OPTIMIZADO)
# ==============================================================================
def modulo_calculo_ttr():
    st.title("🧮 TTR_ARIA - Módulo 3: Cálculo y Valorización TTR")
    st.markdown("Clasificación tarifaria y cálculo de recaudación teórica (TRSUBE) conectada a la Matriz ARIA.")

    st.header("1. Archivos de Entrada")
    col1, col2 = st.columns(2)
    file_tarifario_dmk = col1.file_uploader("1. DMK Resumen_Tarifario_Dominio", type=['xlsx', 'csv'])
    file_matriz_aria = col2.file_uploader("2. Matriz TTR ARIA (Módulo 0)", type=['xlsx'])
    file_nom_ramal = col1.file_uploader("3. Nomenclador Ramal - TS", type=['xlsx'])
    file_ttr_reso = col2.file_uploader("4. Teórica Resoluciones", type=['xlsx'])

    with st.expander("⚙️ Parámetros de Clasificación", expanded=False):
        c1, c2, c3 = st.columns(3)
        anio = c1.number_input("Año de Liquidación", value=2026)
        resolucion = c2.text_input("Número de Resolución", "48")
        filtro_tarifa_vieja = c3.number_input("Corte Tarifa Vieja (Ej: 728.23)", value=728.23)

    if st.button("🚀 Ejecutar Valorización TTR", type="primary", use_container_width=True):
        if not all([file_tarifario_dmk, file_matriz_aria, file_nom_ramal, file_ttr_reso]):
            st.error("⚠️ Faltan archivos requeridos para la liquidación.")
            return

        with st.spinner("Procesando árbol de decisiones TTR (Optimizando RAM)..."):
            try:
                if file_tarifario_dmk.name.endswith('.csv'): df1 = pd.read_csv(file_tarifario_dmk, sep=';', encoding='utf-8-sig')
                else: df1 = pd.read_excel(file_tarifario_dmk)
                
                nom_ts = pd.read_excel(file_nom_ramal, sheet_name='NOMENCLADOR TS' if 'NOMENCLADOR TS' in pd.ExcelFile(file_nom_ramal).sheet_names else 0)
                ttr_reso = pd.read_excel(file_ttr_reso, sheet_name='TTR')
                ttr_sgii_uma2 = pd.read_excel(file_ttr_reso, sheet_name='SGII-UMA2')
                df_aria = pd.read_excel(file_matriz_aria)

                c_concat_aria = buscar_col(df_aria, 'CONCAT')
                col_mes_inf = [c for c in df_aria.columns if c not in [c_concat_aria, 'TS', 'Seccion', 'KM', 'Nominalizacion'] and not str(c).endswith('_Sup') and not str(c).strip() == ''][0]
                col_mes_sup = f"{col_mes_inf}_Sup"
                if col_mes_sup not in df_aria.columns: col_mes_sup = [c for c in df_aria.columns if str(c).strip() == ''][0]

                limites = {}
                for _, row in df_aria.iterrows(): limites[row[c_concat_aria]] = (row[col_mes_inf], row[col_mes_sup])
                def get_lim(k): return limites.get(k, (0.0, 0.0))

                tarifas_1 = {
                    '1SCN': get_lim('1SCN'), '2SCN': get_lim('2SCN'), '3SCN': get_lim('3SCN'), '4SCN': get_lim('4SCN'), '5SCN': get_lim('5SCN'),
                    '1SEN': get_lim('1SEN'), '2SEN': get_lim('2SEN'), '3SEN': get_lim('3SEN'), '4SEN': get_lim('4SEN'), '5SEN': get_lim('5SEN'),
                    '1SEAN': get_lim('1SEAN'), '2SEAN': get_lim('2SEAN'), '3SEAN': get_lim('3SEAN'), '4SEAN': get_lim('4SEAN'), '5SEAN': get_lim('5SEAN')
                }
                tarifas_2 = {
                    '1SCSN': get_lim('1SCSN'), '2SCSN': get_lim('2SCSN'), '3SCSN': get_lim('3SCSN'), '4SCSN': get_lim('4SCSN'), '5SCSN': get_lim('5SCSN'),
                    '1SESN': get_lim('1SESN'), '2SESN': get_lim('2SESN'), '3SESN': get_lim('3SESN'), '4SESN': get_lim('4SESN'), '5SESN': get_lim('5SESN'),
                    '1SEASN': get_lim('1SEASN'), '2SEASN': get_lim('2SEASN'), '3SEASN': get_lim('3SEASN'), '4SEASN': get_lim('4SEASN'), '5SEASN': get_lim('5SEASN')
                }
                tarifas_3 = {'1-4KMCN': (*get_lim('1-4KMCN'), "C"), '1-4KMEN': (*get_lim('1-4KMEN'), "E"), '1-4KMEAN': (*get_lim('1-4KMEAN'), "EA")}
                tarifas_4 = {'1-4KMCSN': (*get_lim('1-4KMCSN'), "C"), '1-4KMESN': (*get_lim('1-4KMESN'), "E"), '1-4KMEASN': (*get_lim('1-4KMEASN'), "EA")}
                
                tarifas_7 = {
                    '5KPCN': (*get_lim('5KPCN'), "C", 5), '6KPCN': (*get_lim('6KPCN'), "C", 6), '7KPCN': (*get_lim('7KPCN'), "C", 7), '8KPCN': (*get_lim('8KPCN'), "C", 8), '9KPCN': (*get_lim('9KPCN'), "C", 9),
                    '5KPEN': (*get_lim('5KPEN'), "E", 5), '6KPEN': (*get_lim('6KPEN'), "E", 6), '7KPEN': (*get_lim('7KPEN'), "E", 7), '8KPEN': (*get_lim('8KPEN'), "E", 8), '9KPEN': (*get_lim('9KPEN'), "E", 9),
                    '5KPEAN': (*get_lim('5KPEAN'), "EA", 5), '6KPEAN': (*get_lim('6KPEAN'), "EA", 6), '7KPEAN': (*get_lim('7KPEAN'), "EA", 7), '8KPEAN': (*get_lim('8KPEAN'), "EA", 8), '9KPEAN': (*get_lim('9KPEAN'), "EA", 9)
                }
                tarifas_8 = {
                    '5KPCSN': (*get_lim('5KPCSN'), "C", 5), '6KPCSN': (*get_lim('6KPCSN'), "C", 6), '7KPCSN': (*get_lim('7KPCSN'), "C", 7), '8KPCSN': (*get_lim('8KPCSN'), "C", 8), '9KPCSN': (*get_lim('9KPCSN'), "C", 9),
                    '5KPESN': (*get_lim('5KPESN'), "E", 5), '6KPESN': (*get_lim('6KPESN'), "E", 6), '7KPESN': (*get_lim('7KPESN'), "E", 7), '8KPESN': (*get_lim('8KPESN'), "E", 8), '9KPESN': (*get_lim('9KPESN'), "E", 9),
                    '5KPEASN': (*get_lim('5KPEASN'), "EA", 5), '6KPEASN': (*get_lim('6KPEASN'), "EA", 6), '7KPEASN': (*get_lim('7KPEASN'), "EA", 7), '8KPEASN': (*get_lim('8KPEASN'), "EA", 8), '9KPEASN': (*get_lim('9KPEASN'), "EA", 9)
                }
                tarifas_cn = {'1-4KMCN2': get_lim('1-4KMCN2')}
                tarifas_en = {'1-4KMEN2': get_lim('1-4KMEN2')}
                tarifas_ean = {'1-4KMEAN2': get_lim('1-4KMEAN2')}
                tarifas_cn_sn = {'1-4KMCSN2': get_lim('1-4KMCSN2')}
                tarifas_en_sn = {'1-4KMESN2': get_lim('1-4KMESN2')}
                tarifas_ean_sn = {'1-4KMEASN2': get_lim('1-4KMEASN2')}

                var_input = ['PROVINCIA', 'MUNICIPIO', 'GRUPO_TARIFARIO', 'AMBA', 'LINEA_SILAS_DNGFF', 'ID_LINEA', 'RAMAL', 'CONTRATO', 'TARIFA', 'DEBITADO', 'DOMINIO', 'ENERGIA', 'USOS', 'RECAUDACION', 'TOTAL DESC POR INTEGRACION', 'DESCUENTO_ATRIBUTOS', 'DESCUENTO_TOTAL']
                df2 = pd.DataFrame()
                for c in var_input:
                    col_real = buscar_col(df1, c)
                    df2[c] = df1[col_real] if col_real else pd.Series(dtype='object')
                
                c_gt = 'GRUPO_TARIFARIO'
                _df2_ = df2[df2[c_gt].isin(["SGI", "SGII", "SGIKM"])].copy() if c_gt else df2.copy()
                
                del df1, df2
                gc.collect()

                c_usos = buscar_col(_df2_, 'USOS')
                c_tar = buscar_col(_df2_, 'TARIFA')
                _df2_[c_usos] = pd.to_numeric(_df2_[c_usos].astype(str).replace({',': ''}, regex=True), errors='coerce').fillna(0) if c_usos else 0
                _df2_[c_tar] = pd.to_numeric(_df2_[c_tar].astype(str).replace({',': ''}, regex=True), errors='coerce').fillna(0).round(2) if c_tar else 0
                
                c_ramal_nom = buscar_col(nom_ts, 'IdRamalNS', 'RAMAL')
                c_ts_nom = buscar_col(nom_ts, 'TIPO DE SERVICIO FINAL', 'SERVICIO', 'TS')
                c_ramal_df = buscar_col(_df2_, 'RAMAL')

                if c_ramal_nom: nom_ts[c_ramal_nom] = nom_ts[c_ramal_nom].astype(str)
                if c_ramal_df: _df2_[c_ramal_df] = _df2_[c_ramal_df].astype(str)
                _df2_ = pd.merge(_df2_, nom_ts[[c_ramal_nom, c_ts_nom]] if c_ramal_nom and c_ts_nom else nom_ts, how='left', left_on=c_ramal_df, right_on=c_ramal_nom)
                
                del nom_ts
                gc.collect()

                c_ts_merge = buscar_col(_df2_, 'TIPO DE SERVICIO FINAL', 'SERVICIO')
                if c_ts_merge: _df2_.rename(columns={c_ts_merge: 'TipoServicio'}, inplace=True)
                _df2_['TipoServicio2'] = _df2_['TipoServicio'].replace('SR', 'E') if 'TipoServicio' in _df2_.columns else 'C'
                
                c_contrato = buscar_col(_df2_, 'CONTRATO')
                _df2_['sin_nominalizar'] = np.where(_df2_[c_contrato] == 627, 1, 0) if c_contrato else 0
                _df2_['PASES'] = np.where((_df2_[c_tar] >= 0) & (_df2_[c_tar] <= 0.5), 1, 0)
                _df2_['FILTRO_1'] = np.where((_df2_[c_tar] < filtro_tarifa_vieja) & (_df2_[c_tar] > 0.5), 1, 0)

                for col, (lim_inf, lim_sup) in tarifas_1.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] <= lim_sup) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 0) & (_df2_['TipoServicio'] != "SR"), 1, 0)
                for col, (lim_inf, lim_sup) in tarifas_2.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] <= lim_sup) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 1) & (_df2_['TipoServicio'] != "SR"), 1, 0)
                for col, (lim_inf, lim_sup, ts_val) in tarifas_3.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] <= lim_sup) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 0) & (_df2_['TipoServicio2'] == ts_val), 1, 0)
                for col, (lim_inf, lim_sup, ts_val) in tarifas_4.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] <= lim_sup) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 1) & (_df2_['TipoServicio2'] == ts_val), 1, 0)
                for col, (lim_inf, lim_sup, ts_val, val_asig) in tarifas_7.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] < lim_sup - 0.5) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 0) & (_df2_['TipoServicio2'] == ts_val), val_asig, 0)
                for col, (lim_inf, lim_sup, ts_val, val_asig) in tarifas_8.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] < lim_sup - 0.5) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 1) & (_df2_['TipoServicio2'] == ts_val), val_asig, 0)

                c_sn_cn = ['1SCN', '2SCN', '3SCN', '4SCN', '5SCN']
                c_kmn_cn = ['1-4KMCN']
                c_sn_en = ['1SEN', '2SEN', '3SEN', '4SEN', '5SEN']
                c_kmn_en = ['1-4KMEN']
                c_kpn_en = ['5KPCN', '6KPCN', '7KPCN', '8KPCN', '9KPCN']
                c_sn_ean = ['1SEAN', '2SEAN', '3SEAN', '4SEAN', '5SEAN']
                c_kmn_ean = ['1-4KMEAN']
                c_kpn_ean = c_kpn_en + ['5KPEN', '6KPEN', '7KPEN', '8KPEN', '9KPEN']
                
                for col, (lim_inf, lim_sup) in tarifas_cn.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] <= lim_sup - 0.5) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 0) & (_df2_[c_sn_cn].sum(axis=1) == 0) & (_df2_[c_kmn_cn].sum(axis=1) == 0) & (_df2_[c_gt] != "DF"), 1, 0)
                for col, (lim_inf, lim_sup) in tarifas_en.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] <= lim_sup - 0.5) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 0) & (_df2_[c_sn_en].sum(axis=1) == 0) & (_df2_[c_kmn_en].sum(axis=1) == 0) & (_df2_[c_kpn_en].sum(axis=1) == 0) & (_df2_[c_gt] != "DF"), 1, 0)
                for col, (lim_inf, lim_sup) in tarifas_ean.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] <= lim_sup - 0.5) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 0) & (_df2_[c_sn_ean].sum(axis=1) == 0) & (_df2_[c_kmn_ean].sum(axis=1) == 0) & (_df2_[c_kpn_ean].sum(axis=1) == 0) & (_df2_[c_gt] != "DF"), 1, 0)
                
                c_sn_cSn = ['1SCSN', '2SCSN', '3SCSN', '4SCSN', '5SCSN']
                c_kmn_cSn = ['1-4KMCSN']
                c_sn_eSn = ['1SESN', '2SESN', '3SESN', '4SESN', '5SESN']
                c_kmn_eSn = ['1-4KMESN']
                c_kpn_eSn = ['5KPCSN', '6KPCSN', '7KPCSN', '8KPCSN', '9KPCSN']
                c_sn_eaSn = ['1SEASN', '2SEASN', '3SEASN', '4SEASN', '5SEASN']
                c_kmn_eaSn = ['1-4KMEASN']
                c_kpn_eaSn = c_kpn_eSn + ['5KPESN', '6KPESN', '7KPESN', '8KPESN', '9KPESN']

                for col, (lim_inf, lim_sup) in tarifas_cn_sn.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] <= lim_sup - 0.5) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 1) & (_df2_[c_sn_cSn].sum(axis=1) == 0) & (_df2_[c_kmn_cSn].sum(axis=1) == 0) & (_df2_[c_gt] != "DF"), 1, 0)
                for col, (lim_inf, lim_sup) in tarifas_en_sn.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] <= lim_sup - 0.5) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 1) & (_df2_[c_sn_eSn].sum(axis=1) == 0) & (_df2_[c_kmn_eSn].sum(axis=1) == 0) & (_df2_[c_kpn_eSn].sum(axis=1) == 0) & (_df2_[c_gt] != "DF"), 1, 0)
                for col, (lim_inf, lim_sup) in tarifas_ean_sn.items(): _df2_[col] = np.where((_df2_[c_tar] >= lim_inf - 0.5) & (_df2_[c_tar] <= lim_sup - 0.5) & (_df2_['PASES'] == 0) & (_df2_['sin_nominalizar'] == 1) & (_df2_[c_sn_eaSn].sum(axis=1) == 0) & (_df2_[c_kmn_eaSn].sum(axis=1) == 0) & (_df2_[c_kpn_eaSn].sum(axis=1) == 0) & (_df2_[c_gt] != "DF"), 1, 0)

                _df2_['Filtro1-4KMCN'] = np.where((_df2_['TipoServicio2'] == 'C') & (_df2_[c_sn_en + c_sn_ean].sum(axis=1) != 0) & (_df2_['1-4KMCN2'] == 1), 4, 0)
                _df2_['Filtro1-4KMEN'] = np.where((_df2_['TipoServicio2'] == 'E') & (_df2_[c_sn_ean].sum(axis=1) != 0) & (_df2_['1-4KMEN2'] == 1), 4, 0)
                _df2_['Filtro1-4KMEAN'] = np.where((_df2_['TipoServicio2'] == 'EA') & (_df2_['1-4KMEAN2'] == 1), 4, 0)
                _df2_['Filtro1-4KMCSN'] = np.where((_df2_['TipoServicio2'] == 'C') & (_df2_[c_sn_eSn + c_sn_eaSn].sum(axis=1) != 0) & (_df2_['1-4KMCSN2'] == 1), 4, 0)
                _df2_['Filtro1-4KMESN'] = np.where((_df2_['TipoServicio2'] == 'E') & (_df2_[c_sn_eaSn].sum(axis=1) != 0) & (_df2_['1-4KMESN2'] == 1), 4, 0)
                _df2_['Filtro1-4KMEASN'] = np.where((_df2_['TipoServicio2'] == 'EA') & (_df2_['1-4KMEASN2'] == 1), 4, 0)

                _df2_['seccionada_correcta_1'] = np.select([(_df2_['1-4KMCN2'] == 1) & (_df2_['Filtro1-4KMCN'] != 4) & (_df2_['1SEN'] == 1), (_df2_['1-4KMCN2'] == 1) & (_df2_['Filtro1-4KMCN'] != 4) & (_df2_['2SEN'] == 1), (_df2_['1-4KMCN2'] == 1) & (_df2_['Filtro1-4KMCN'] != 4) & (_df2_['3SEN'] == 1), (_df2_['1-4KMCN2'] == 1) & (_df2_['Filtro1-4KMCN'] != 4) & (_df2_['4SEN'] == 1), (_df2_['1-4KMCN2'] == 1) & (_df2_['Filtro1-4KMCN'] != 4) & (_df2_['5SEN'] == 1)], [1, 2, 3, 4, 5], default=0)
                _df2_['seccionada_correcta_3'] = np.select([(_df2_['Filtro1-4KMCSN'] != 4) & (_df2_['1-4KMCSN2'] == 1) & (_df2_['1SESN'] == 1), (_df2_['Filtro1-4KMCSN'] != 4) & (_df2_['1-4KMCSN2'] == 1) & (_df2_['2SESN'] == 1), (_df2_['Filtro1-4KMCSN'] != 4) & (_df2_['1-4KMCSN2'] == 1) & (_df2_['3SESN'] == 1), (_df2_['Filtro1-4KMCSN'] != 4) & (_df2_['1-4KMCSN2'] == 1) & (_df2_['4SESN'] == 1), (_df2_['Filtro1-4KMCSN'] != 4) & (_df2_['1-4KMCSN2'] == 1) & (_df2_['5SESN'] == 1)], [1, 2, 3, 4, 5], default=0)
                _df2_['seccionada_correcta_2'] = np.select([(_df2_['1-4KMEN2'] == 1) & (_df2_['Filtro1-4KMEN'] != 4) & (_df2_['1SEAN'] == 1), (_df2_['1-4KMEN2'] == 1) & (_df2_['Filtro1-4KMEN'] != 4) & (_df2_['2SEAN'] == 1), (_df2_['1-4KMEN2'] == 1) & (_df2_['Filtro1-4KMEN'] != 4) & (_df2_['3SEAN'] == 1), (_df2_['1-4KMEN2'] == 1) & (_df2_['Filtro1-4KMEN'] != 4) & (_df2_['4SEAN'] == 1), (_df2_['1-4KMEN2'] == 1) & (_df2_['Filtro1-4KMEN'] != 4) & (_df2_['5SEAN'] == 1)], [1, 2, 3, 4, 5], default=0)
                _df2_['seccionada_correcta_4'] = np.select([(_df2_['1-4KMESN2'] == 1) & (_df2_['Filtro1-4KMESN'] != 4) & (_df2_['1SEASN'] == 1), (_df2_['1-4KMESN2'] == 1) & (_df2_['Filtro1-4KMESN'] != 4) & (_df2_['2SEASN'] == 1), (_df2_['1-4KMESN2'] == 1) & (_df2_['Filtro1-4KMESN'] != 4) & (_df2_['3SEASN'] == 1), (_df2_['1-4KMESN2'] == 1) & (_df2_['Filtro1-4KMESN'] != 4) & (_df2_['4SEASN'] == 1), (_df2_['1-4KMESN2'] == 1) & (_df2_['Filtro1-4KMESN'] != 4) & (_df2_['5SEASN'] == 1)], [1, 2, 3, 4, 5], default=0)

                _df2_['sec_c'] = np.where((_df2_[c_sn_cn].sum(axis=1) > 0) | (_df2_[c_sn_cSn].sum(axis=1) > 0) | (_df2_['1-4KMCN'] > 0) | (_df2_['1-4KMCSN'] > 0), 1, 0)
                _df2_['sec_e'] = np.where((_df2_[c_sn_en].sum(axis=1) > 0) | (_df2_[c_sn_eSn].sum(axis=1) > 0) | (_df2_['1-4KMEN'] > 0) | (_df2_['1-4KMESN'] > 0), 1, 0)
                _df2_['sec_ea'] = np.where((_df2_[c_sn_ean].sum(axis=1) > 0) | (_df2_[c_sn_eaSn].sum(axis=1) > 0) | (_df2_['1-4KMEAN'] > 0) | (_df2_['1-4KMEASN'] > 0), 1, 0)

                _df2_['km&p_c'] = np.where((_df2_[c_kpn_en[:5]].sum(axis=1) > 0) | (_df2_[c_kpn_eSn[:5]].sum(axis=1) > 0) | (_df2_['1-4KMCN2'] > 0) | (_df2_['1-4KMCSN2'] > 0), 1, 0)
                _df2_['km&p_e'] = np.where((_df2_[c_kpn_en[5:]].sum(axis=1) > 0) | (_df2_[c_kpn_eSn[5:10]].sum(axis=1) > 0) | (_df2_['1-4KMEN2'] > 0) | (_df2_['1-4KMESN2'] > 0), 1, 0)
                _df2_['km&p_ea'] = np.where((_df2_['5KPEAN'] > 0) | (_df2_['1-4KMEAN2'] > 0) | (_df2_['1-4KMEASN2'] > 0), 1, 0)

                _df2_['compilado_ts'] = np.select([((_df2_[['km&p_c', 'km&p_e', 'km&p_ea']] == 1).any(axis=1)) & (_df2_[['seccionada_correcta_1', 'seccionada_correcta_2', 'seccionada_correcta_3', 'seccionada_correcta_4']].sum(axis=1) == 0), (_df2_['sec_c'] == 1), (_df2_['sec_e'] == 1), (_df2_['sec_ea'] == 1), (_df2_['PASES'] == 1), (_df2_[['sec_c', 'sec_e', 'sec_ea', 'km&p_c', 'km&p_e', 'km&p_ea']].sum(axis=1) == 0)], [_df2_['TipoServicio2'], 'C', 'E', 'EA', _df2_['TipoServicio2'], _df2_['TipoServicio2']], default="S/D")

                _df2_['norm_por_tarifa'] = np.where(((_df2_[c_sn_cn + c_sn_en + c_sn_ean].sum(axis=1) > 0) | (_df2_[['1-4KMCN', '1-4KMEN', '1-4KMEAN']].sum(axis=1) > 0) | (_df2_[c_kpn_en + c_kpn_ean].sum(axis=1) > 0) | (_df2_[['1-4KMCN2', '1-4KMEN2', '1-4KMEAN2']].sum(axis=1) > 0)), "N", np.where(_df2_['FILTRO_1'] == 1, "Tarifa Vieja", np.where(_df2_['PASES'] == 1, "N", "SN")))
                
                _df2_['tarifa_s'] = np.where(((_df2_[c_sn_cn + c_sn_en + c_sn_ean + c_sn_cSn + c_sn_eSn + c_sn_eaSn].sum(axis=1) > 0)) & (_df2_[['Filtro1-4KMCN', 'Filtro1-4KMEN', 'Filtro1-4KMEAN', 'Filtro1-4KMCSN', 'Filtro1-4KMESN', 'Filtro1-4KMEASN']].sum(axis=1) == 0) & (((_df2_['compilado_ts'] == 'C') & (_df2_['sec_c'] == 1)) | ((_df2_['compilado_ts'] == 'E') & (_df2_['sec_e'] == 1)) | ((_df2_['compilado_ts'] == 'EA') & (_df2_['sec_ea'] == 1))), 1, 0)
                _df2_['tarifa_km'] = np.where(((_df2_[['1-4KMCN', '1-4KMEN', '1-4KMEAN', '1-4KMCSN','1-4KMESN', '1-4KMEASN']].sum(axis=1) > 0) | (_df2_[['1-4KMCN2', '1-4KMEN2', '1-4KMEAN2', '1-4KMCSN2', '1-4KMESN2', '1-4KMEASN2']].sum(axis=1) > 0)) & (_df2_[['seccionada_correcta_1', 'seccionada_correcta_3', 'seccionada_correcta_2', 'seccionada_correcta_4']].sum(axis=1) == 0), 1, 0)
                _df2_['tarifa_kp'] = np.where((_df2_[c_kpn_en + c_kpn_ean + c_kpn_eSn + c_kpn_eaSn].sum(axis=1) > 0), 1, 0)
                _df2_['tarifa_PASE'] = np.where((_df2_['PASES'] == 1), 1, 0)

                _df2_['compilado_tt'] = np.where(_df2_['tarifa_s'] != 0, 'S', np.where(_df2_['tarifa_km'] != 0, 'KM', np.where(_df2_['tarifa_kp'] != 0, 'KP', np.where(_df2_['tarifa_PASE'] != 0, 'P', np.where((_df2_['seccionada_correcta_1'] + _df2_['seccionada_correcta_2'] + _df2_['seccionada_correcta_3'] + _df2_['seccionada_correcta_4']) != 0, 'S', 'S/D')))))

                _df2_['sec_1'] = np.where((_df2_[['1SCN', '1SCSN', '1SEN', '1SESN', '1SEAN', '1SEASN']].sum(axis=1) > 0) & (_df2_[['Filtro1-4KMCN', 'Filtro1-4KMEN', 'Filtro1-4KMEAN', 'Filtro1-4KMCSN', 'Filtro1-4KMESN', 'Filtro1-4KMEASN']].sum(axis=1) == 0), 1, 0)
                _df2_['sec_2'] = np.where((_df2_[['2SCN', '2SCSN', '2SEN', '2SESN', '2SEAN', '2SEASN']].sum(axis=1) > 0) & (_df2_[['Filtro1-4KMCN', 'Filtro1-4KMEN', 'Filtro1-4KMEAN', 'Filtro1-4KMCSN', 'Filtro1-4KMESN', 'Filtro1-4KMEASN']].sum(axis=1) == 0), 2, 0)
                _df2_['sec_3'] = np.where((_df2_[['3SCN', '3SCSN', '3SEN', '3SESN', '3SEAN', '3SEASN']].sum(axis=1) > 0) & (_df2_[['Filtro1-4KMCN', 'Filtro1-4KMEN', 'Filtro1-4KMEAN', 'Filtro1-4KMCSN', 'Filtro1-4KMESN', 'Filtro1-4KMEASN']].sum(axis=1) == 0), 3, 0)
                _df2_['sec_4'] = np.where((_df2_[['4SCN', '4SCSN', '4SEN', '4SESN', '4SEAN', '4SEASN']].sum(axis=1) > 0) & (_df2_[['Filtro1-4KMCN', 'Filtro1-4KMEN', 'Filtro1-4KMEAN', 'Filtro1-4KMCSN', 'Filtro1-4KMESN', 'Filtro1-4KMEASN']].sum(axis=1) == 0), 4, 0)
                _df2_['sec_5'] = np.where((_df2_[['5SCN', '5SCSN', '5SEN', '5SESN', '5SEAN', '5SEASN']].sum(axis=1) > 0) & (_df2_[['Filtro1-4KMCN', 'Filtro1-4KMEN', 'Filtro1-4KMEAN', 'Filtro1-4KMCSN', 'Filtro1-4KMESN', 'Filtro1-4KMEASN']].sum(axis=1) == 0), 5, 0)

                _df2_['seccionadas_final'] = np.where(_df2_['PASES'] == 1, 1, _df2_[['sec_1', 'sec_2', 'sec_3', 'sec_4', 'sec_5']].sum(axis=1))
                _df2_['sec_1_4'] = np.where((_df2_[['1-4KMCN', '1-4KMEN', '1-4KMEAN', '1-4KMCSN', '1-4KMESN', '1-4KMEASN']].sum(axis=1) > 0) | (_df2_[['1-4KMCN2', '1-4KMEN2', '1-4KMEAN2', '1-4KMCSN2', '1-4KMESN2', '1-4KMEASN2']].sum(axis=1) > 0) & (_df2_[['seccionada_correcta_1', 'seccionada_correcta_2', 'seccionada_correcta_3', 'seccionada_correcta_4']].eq(0).all(axis=1)), 4, 0)

                _df2_['kilometricas'] = np.where((_df2_['compilado_ts'] == 'C') & (_df2_['norm_por_tarifa'] == 'N'), _df2_[c_kpn_en[:5]].sum(axis=1), np.where((_df2_['compilado_ts'] == 'E') & (_df2_['norm_por_tarifa'] == 'N'), _df2_[c_kpn_en[5:]].sum(axis=1), np.where((_df2_['compilado_ts'] == 'EA') & (_df2_['norm_por_tarifa'] == 'N'), _df2_[['5KPEAN', '6KPEAN', '7KPEAN', '8KPEAN', '9KPEAN']].sum(axis=1), np.where((_df2_['compilado_ts'] == 'C') & (_df2_['norm_por_tarifa'] == 'SN'), _df2_[c_kpn_eSn[:5]].sum(axis=1), np.where((_df2_['compilado_ts'] == 'E') & (_df2_['norm_por_tarifa'] == 'SN'), _df2_[c_kpn_eSn[5:10]].sum(axis=1), _df2_[['5KPEASN', '6KPEASN', '7KPEASN', '8KPEASN', '9KPEASN']].sum(axis=1))))))
                _df2_['kilometricas_por_TS'] = np.where((_df2_['TipoServicio'] == 'C') & (_df2_['sin_nominalizar'] == 0), _df2_[c_kpn_en[:5]].replace(0, np.nan).min(axis=1).fillna(0), np.where((_df2_['TipoServicio'] == 'E') & (_df2_['sin_nominalizar'] == 0), _df2_[c_kpn_en[5:]].replace(0, np.nan).min(axis=1).fillna(0), np.where((_df2_['TipoServicio'] == 'EA') & (_df2_['sin_nominalizar'] == 0), _df2_[['5KPEAN', '6KPEAN', '7KPEAN', '8KPEAN', '9KPEAN']].replace(0, np.nan).min(axis=1).fillna(0), np.where((_df2_['TipoServicio'] == 'C') & (_df2_['sin_nominalizar'] == 1), _df2_[c_kpn_eSn[:5]].replace(0, np.nan).min(axis=1).fillna(0), np.where((_df2_['TipoServicio'] == 'E') & (_df2_['sin_nominalizar'] == 1), _df2_[c_kpn_eSn[5:10]].replace(0, np.nan).min(axis=1).fillna(0), np.where((_df2_['TipoServicio'] == 'EA') & (_df2_['sin_nominalizar'] == 1), _df2_[['5KPEASN', '6KPEASN', '7KPEASN', '8KPEASN', '9KPEASN']].replace(0, np.nan).min(axis=1).fillna(0), 0))))))

                _df2_['compilado_seccion'] = np.where(_df2_['compilado_tt'] == "S", _df2_['seccionadas_final'], np.where(_df2_['compilado_tt'] == "P", 1, np.where(_df2_['compilado_tt'] == "KM", _df2_['sec_1_4'], np.where(_df2_['compilado_tt'] == "KP", _df2_['kilometricas_por_TS'], 0))))
                _df2_['final_seccion'] = np.where((_df2_[c_gt] == "SGII") & (_df2_['compilado_seccion'].isin([1, 2, 3])), 4, _df2_['compilado_seccion']) if c_gt else _df2_['compilado_seccion']

                _df2_['Año'] = anio
                _df2_['Resolucion'] = resolucion
                _df2_['CONCAT_MACHEO2'] = _df2_['Año'].astype(int).astype(str) + _df2_['Resolucion'].astype(str) + _df2_['final_seccion'].astype(int).astype(str) + (_df2_[c_gt].astype(str) if c_gt else "") + _df2_['compilado_ts'].astype(str) + _df2_['norm_por_tarifa'].astype(str)
                c_id = buscar_col(_df2_, 'ID_LINEA')
                _df2_['CONCAT_MACHEO3'] = _df2_['Año'].astype(int).astype(str) + _df2_['Resolucion'].astype(str) + _df2_['final_seccion'].astype(int).astype(str) + (_df2_[c_gt].astype(str) if c_gt else "") + (_df2_[c_id].astype(str) if c_id else "") + _df2_['compilado_ts'].astype(str) + _df2_['norm_por_tarifa'].astype(str)

                c_concat_reso = buscar_col(ttr_reso, 'CONCAT')
                c_ttr_ec_reso = buscar_col(ttr_reso, 'TTR E.C.', 'TTR')
                _df2_ = pd.merge(_df2_, ttr_reso[[c_concat_reso, c_ttr_ec_reso]] if c_concat_reso and c_ttr_ec_reso else ttr_reso, how='left', left_on='CONCAT_MACHEO2', right_on=c_concat_reso if c_concat_reso else 'CONCAT').fillna({c_ttr_ec_reso if c_ttr_ec_reso else 'TTR E.C.': 0})
                _df2_ = _df2_.rename(columns={c_ttr_ec_reso if c_ttr_ec_reso else "TTR E.C.": "Tarifa TRSUBE"})
                if c_concat_reso and c_concat_reso in _df2_.columns: _df2_ = _df2_.drop(columns=[c_concat_reso])

                c_concat_uma = buscar_col(ttr_sgii_uma2, 'CONCAT')
                c_ttr_ec_uma = buscar_col(ttr_sgii_uma2, 'TTR E.C.', 'TTR')
                _df2_ = pd.merge(_df2_, ttr_sgii_uma2[[c_concat_uma, c_ttr_ec_uma]] if c_concat_uma and c_ttr_ec_uma else ttr_sgii_uma2, how='left', left_on='CONCAT_MACHEO3', right_on=c_concat_uma if c_concat_uma else 'CONCAT').fillna({c_ttr_ec_uma if c_ttr_ec_uma else 'TTR E.C.': 0})
                _df2_ = _df2_.rename(columns={c_ttr_ec_uma if c_ttr_ec_uma else "TTR E.C.": "Tarifa TRSUBE2"})
                if c_concat_uma and c_concat_uma in _df2_.columns: _df2_ = _df2_.drop(columns=[c_concat_uma])

                _df2_['Tarifa TRSUBE_FINAL'] = np.where(_df2_['Tarifa TRSUBE2'] == 0, _df2_['Tarifa TRSUBE'], _df2_['Tarifa TRSUBE2'])
                _df2_['Recaudacion_TRSUBE'] = _df2_['Tarifa TRSUBE_FINAL'] * _df2_[c_usos]

                c_ene = buscar_col(_df2_, 'ENERGIA')
                if c_ene:
                    condiciones = [_df2_[c_ene] == 1, _df2_[c_ene] == 2, _df2_[c_ene] == 3]
                    _df2_['Recaudacion_TRSUBE'] = _df2_['Recaudacion_TRSUBE'] * np.select(condiciones, [1.3, 1.5, 1.0], default=1)

                _df2_['SubSeccion'] = None
                def asignar_subsecciones(df, dict_t, flt_sn):
                    for _, (l_inf, l_sup, t_srv, sec) in dict_t.items():
                        s_rng = np.linspace(l_inf, l_sup, 4)
                        for i in range(3):
                            msk = ((df[c_tar] >= s_rng[i] - 0.5) & (df[c_tar] < s_rng[i+1] - 0.5) & (df['PASES'] == 0) & (df['sin_nominalizar'] == flt_sn) & (df['TipoServicio2'] == t_srv))
                            df.loc[msk, 'SubSeccion'] = f"{sec}-{i+1}"

                asignar_subsecciones(_df2_, tarifas_7, 0)
                asignar_subsecciones(_df2_, tarifas_8, 1)
                _df2_['SubSeccion'] = _df2_['SubSeccion'].fillna(_df2_['final_seccion'].astype(str))

                buf_salida = io.BytesIO()
                _df2_.to_excel(buf_salida, index=False, sheet_name='Valorizacion_TTR')
                buf_salida.seek(0)
                
                del df_aria, _df2_
                gc.collect()

                st.success("✅ ¡Valorización TTR calculada con éxito! Matrices y diccionarios cruzados.")
                st.download_button(label="📥 Descargar Base Macheo TTR (.xlsx)", data=buf_salida, file_name="macheo_ttr_ARIA.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            except Exception as e: st.error(f"Error procesando TTR: {e}")

# ==============================================================================
# CONTROLADOR LATERAL
# ==============================================================================
st.sidebar.image("https://cdn-icons-png.flaticon.com/512/1792/1792404.png", width=60)
st.sidebar.title("Menú TTR_ARIA")
modulo_seleccionado = st.sidebar.radio("Navegación", ["Módulo 0: Tarifas JN", "Módulo 1: Liquidación DMK", "Módulo 3: Cálculo TTR"])
st.sidebar.markdown("---")
st.sidebar.info("Proyecto ARIA v2.8\n\nMotor unificado de cálculos TTR.")

if modulo_seleccionado == "Módulo 0: Tarifas JN": modulo_tarifas()
elif modulo_seleccionado == "Módulo 1: Liquidación DMK": modulo_dmk()
else: modulo_calculo_ttr()
